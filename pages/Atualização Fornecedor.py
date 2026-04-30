import streamlit as st
import pandas as pd
import openpyxl
import os
from datetime import date

st.set_page_config(layout="wide", page_title="QSMS - Cadastro")

FILE_PATH = r'C:\Users\gabriel.silva\OneDrive - compasa.com.br\QSMS - Administrativo - Área de Cadastros\Controle Cadastros.xlsx'
NOME_ABA = 'Alt_Att Fornec'

def formata_zerados():
    if st.session_state.codigo_input:
        st.session_state.codigo_input = st.session_state.codigo_input.zfill(6)
    if st.session_state.loja_input:
        st.session_state.loja_input = st.session_state.loja_input.zfill(2)

def limpar_tudo():
    for key in list(st.session_state.keys()):
        if key == 'atualizacao_input':
            st.session_state[key] = None
        elif key == 'check_branco':
            st.session_state[key] = False
        # Não apaga a data, a flag de limpar, a busca, nem a flag de sucesso
        elif key in ['data_input', 'deve_limpar', 'busca_fornecedor', 'mostrar_sucesso']:
            pass 
        else:
            st.session_state[key] = ""

# Flag para limpar cache sem bugar a tela
if st.session_state.get('deve_limpar', False):
    limpar_tudo()
    st.session_state['deve_limpar'] = False

# --- FUNÇÃO DO POP-UP (MODAL) ---
@st.dialog("📝 Cadastrar Nova Atualização", width='medium')
def modal_novo_registro():
    c1, c2, c3, c4 = st.columns([1, 0.8, 1.5, 2])
    codigo = c1.text_input("Código", max_chars=6, key="codigo_input", on_change=formata_zerados)
    loja = c2.text_input("Loja", max_chars=2, key="loja_input", on_change=formata_zerados)
    
    data_solic = c3.date_input("Data Solicitação", date.today(), format="DD/MM/YYYY", key="data_input")
    
    atualizacao = c4.selectbox("O que será atualizado?", 
                             ["Dados Bancários", "Dados Cadastrais"], 
                             index=None, 
                             placeholder="Selecione uma opção...",
                             key="atualizacao_input")

    c5, c6 = st.columns([3, 1])
    razao_social = c5.text_input("Razão Social", key="razao_input")
    solicitante = c6.text_input("Solicitante", key="solicitante_input")

    if atualizacao:
        st.write("---")
        
        dados_em_branco = st.checkbox("Dados em branco", key="check_branco")

        dados_antigos_val = ""
        dados_novos_val = ""

        if atualizacao == "Dados Bancários":
            st.markdown("**Dados Bancários ANTIGOS:**")
            if dados_em_branco:
                st.info("Campo preenchido automaticamente como 'Em branco'.")
                dados_antigos_val = "Em branco"
            else:
                ant_col1, ant_col2, ant_col3 = st.columns(3)
                b_ant = ant_col1.text_input("Banco", key="b_ant")
                a_ant = ant_col2.text_input("Agência", key="a_ant")
                c_ant = ant_col3.text_input("Conta", key="c_ant")
                dados_antigos_val = f"Banco: {b_ant} - Ag: {a_ant} - C/C: {c_ant}"
            st.write("---")

            st.markdown("**Dados Bancários NOVOS:**")
            nov_col1, nov_col2, nov_col3 = st.columns(3)
            b_nov = nov_col1.text_input("Banco", key="b_nov")
            a_nov = nov_col2.text_input("Agência", key="a_nov")
            c_nov = nov_col3.text_input("Conta", key="c_nov")
            dados_novos_val = f"Banco: {b_nov} - Ag: {a_nov} - C/C: {c_nov}"
            st.write("---")

        elif atualizacao == "Dados Cadastrais":
            st.markdown("**Dados Cadastrais:**")
            if dados_em_branco:
                st.info("Campo preenchido automaticamente como 'Em branco'.")
                dados_antigos_val = "Em branco"
            else:
                dados_antigos_val = st.text_area("Dados Antigos (Cadastrais)", height=100, key="cad_antigo")
                
            dados_novos_val = st.text_area("Dados Novos (Cadastrais)", height=100, key="cad_novo")

        obs = st.text_input("Observação", key="obs_input")

        st.write("---")
        
        # --- NOVOS CHECKBOXES ---
        col_chk1, col_chk2 = st.columns(2)
        with col_chk1:
            auth_pagamento = st.checkbox("Possui autorização de pagamento?", key="chk_auth")
        with col_chk2:
            fornecedor_novo = st.checkbox("É um fornecedor novo?", key="chk_novo")

        st.write("---")
        
        # --- CAMPO DE SENHA AQUI ---
        senha_validacao = st.text_input("🔑 Senha de Autorização para Salvar", type="password", key="senha_val")

        col_btn1, col_btn2 = st.columns([1, 5])
        with col_btn1:
            submit = st.button("💾 Salvar", width='content')
        with col_btn2:
            if st.button("🧹 Limpar Campos"):
                st.session_state['deve_limpar'] = True
                st.rerun()

        if submit:
            # --- VALIDAÇÃO DA SENHA E REGRAS DE NEGÓCIO ---
            if senha_validacao != "QSMS2026":
                st.error("Senha incorreta! Você não tem autorização para salvar esta atualização.")
            elif auth_pagamento and not obs.strip():
                st.warning("Como possui autorização de pagamento, o campo 'Observação' é obrigatório!")
            else:
                codigo_final = st.session_state.codigo_input
                loja_final = st.session_state.loja_input

                if not codigo_final or not loja_final or not razao_social:
                    st.warning("Preencha o Código, a Loja e a Razão Social para registrar na planilha!")
                else:
                    try:
                        wb = openpyxl.load_workbook(FILE_PATH)
                        ws = wb[NOME_ABA]

                        next_row = 1
                        for r in range(ws.max_row, 0, -1):
                            if ws.cell(row=r, column=1).value is not None:
                                next_row = r + 1
                                break

                        # Tratamento de Sim/Não para as novas colunas
                        val_auth = "Sim" if auth_pagamento else "Não"
                        val_novo = "Sim" if fornecedor_novo else "Não"

                        ws.cell(row=next_row, column=1, value=codigo_final)
                        ws.cell(row=next_row, column=2, value=loja_final)
                        ws.cell(row=next_row, column=3, value=razao_social)
                        ws.cell(row=next_row, column=4, value=data_solic.strftime('%d/%m/%Y'))
                        ws.cell(row=next_row, column=5, value=atualizacao)
                        ws.cell(row=next_row, column=6, value=dados_antigos_val)
                        ws.cell(row=next_row, column=7, value=dados_novos_val)
                        ws.cell(row=next_row, column=8, value=solicitante)
                        ws.cell(row=next_row, column=9, value=obs)
                        ws.cell(row=next_row, column=10, value=val_auth) # Coluna J
                        ws.cell(row=next_row, column=11, value=val_novo) # Coluna K

                        wb.save(FILE_PATH)
                        
                        st.session_state['mostrar_sucesso'] = True
                        st.session_state['deve_limpar'] = True
                        st.rerun()
                        
                    except PermissionError:
                        st.error("O arquivo Excel está aberto. Feche-o antes de salvar.")
                    except Exception as e:
                        st.error(f"Erro ao salvar: {e}")

# --- TELA PRINCIPAL ---
def main():
    st.title("Atualização de Fornecedores")

    if st.session_state.get('mostrar_sucesso', False):
        st.success("Tudo certo! Registro inserido na planilha com sucesso.")
        st.session_state['mostrar_sucesso'] = False 

    st.divider()

    col_hist, col_btn = st.columns([4, 1])
    with col_hist:
        st.subheader("Histórico de Atualizações")
    with col_btn:
        if st.button("➕ Novo Registro", width='stretch', type="primary"):
            limpar_tudo() # <-- Puxa a descarga aqui antes de abrir a tela
            modal_novo_registro()

    termo_busca = st.text_input("🔍 Buscar por Código ou Razão Social:", key="busca_fornecedor")

    try:
        df_view = pd.read_excel(FILE_PATH, sheet_name=NOME_ABA, header=1)
        df_view.columns = df_view.columns.str.strip()
        
        if 'Data Solicitação' in df_view.columns:
            df_view['Data Solicitação'] = pd.to_datetime(df_view['Data Solicitação'], errors='coerce').dt.strftime('%d/%m/%Y')
        
        if 'Código' in df_view.columns:
            df_view['Código'] = df_view['Código'].apply(lambda x: str(x).replace('.0', '').zfill(6) if pd.notnull(x) and str(x) != 'nan' else '')
        if 'Loja' in df_view.columns:
            df_view['Loja'] = df_view['Loja'].apply(lambda x: str(x).replace('.0', '').zfill(2) if pd.notnull(x) and str(x) != 'nan' else '')

        if termo_busca:
            termo = termo_busca.lower()
            df_view = df_view[
                df_view['Código'].astype(str).str.lower().str.contains(termo, na=False) |
                df_view['Razão Social'].astype(str).str.lower().str.contains(termo, na=False)
            ]

        df_view = df_view.iloc[::-1]

        st.dataframe(df_view, width='stretch')
    except Exception as e:
        st.error(f"Erro ao carregar a visualização: {e}")

if __name__ == "__main__":
    main()